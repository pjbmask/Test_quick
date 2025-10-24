import openpyxl
from pathlib import Path
import shutil

# 폴더 경로 설정
folder_path = Path(r"C:\py\Test_quick\reference")

# 파일 경로
base_template = folder_path / "Base_testsheet.xlsx"
data_source = folder_path / "base_통제활동_필요증빙명_교체.xlsx"

# 데이터 소스 파일 읽기
wb_data = openpyxl.load_workbook(data_source)
ws_data = wb_data.active

# 헤더를 제외한 데이터 행 처리 (2행부터 시작)
for row in ws_data.iter_rows(min_row=2, values_only=False):
    통제코드 = row[0].value  # A열
    통제활동 = row[1].value  # B열
    필요증빙명 = row[2].value  # C열
    
    # 값이 없으면 스킵
    if not 통제코드:
        continue
    
    # 새 파일명 생성
    new_filename = folder_path / f"{통제코드}.xlsx"
    
    # Base_testsheet 복사
    shutil.copy2(base_template, new_filename)
    
    # 복사된 파일 열기
    wb_new = openpyxl.load_workbook(new_filename)
    ws_new = wb_new.active
    
    # B7에 통제활동, N7에 필요증빙명 입력
    ws_new['B7'] = 통제활동
    ws_new['N7'] = 필요증빙명
    
    # 저장
    wb_new.save(new_filename)
    wb_new.close()
    
    print(f"생성 완료: {new_filename.name}")

wb_data.close()
print("\n모든 파일 생성이 완료되었습니다.")
