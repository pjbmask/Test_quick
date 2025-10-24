import openpyxl
from copy import copy
import os

def copy_sheet_to_files(template_file, file_list_excel, template_sheet_name=None):
    """
    템플릿 시트를 파일 목록의 각 파일에 마지막 시트로 추가
    
    Args:
        template_file: 템플릿 시트가 있는 엑셀 파일 경로
        file_list_excel: 파일 목록이 A열에 있는 엑셀 파일 경로
        template_sheet_name: 복사할 시트 이름 (None이면 첫 번째 시트)
    """
    
    # 1. 파일 목록 읽기
    print("파일 목록을 읽는 중...")
    list_wb = openpyxl.load_workbook(file_list_excel)
    list_ws = list_wb.active
    
    file_paths = []
    for row in range(2, list_ws.max_row + 1):  # 1행은 헤더라고 가정
        file_path = list_ws[f'A{row}'].value
        if file_path and os.path.exists(str(file_path)):
            file_paths.append(str(file_path))
        elif file_path:
            print(f"경고: 파일을 찾을 수 없음 - {file_path}")
    
    print(f"총 {len(file_paths)}개 파일 발견\n")
    
    # 2. 템플릿 파일 로드
    print("템플릿 시트를 로드하는 중...")
    template_wb = openpyxl.load_workbook(template_file)
    
    if template_sheet_name:
        template_sheet = template_wb[template_sheet_name]
    else:
        template_sheet = template_wb.active
    
    print(f"템플릿 시트: '{template_sheet.title}'\n")
    
    # 3. 각 파일에 시트 추가
    success_count = 0
    fail_count = 0
    
    for idx, target_file in enumerate(file_paths, 1):
        try:
            print(f"[{idx}/{len(file_paths)}] 처리 중: {os.path.basename(target_file)}")
            
            # 대상 파일 열기
            target_wb = openpyxl.load_workbook(target_file)
            
            # 동일한 이름의 시트가 있는지 확인
            if template_sheet.title in target_wb.sheetnames:
                print(f"  - 기존 '{template_sheet.title}' 시트 삭제")
                del target_wb[template_sheet.title]
            
            # 시트 통째로 복사 (workbook의 copy_worksheet 메서드 사용)
            new_sheet = target_wb.copy_worksheet(template_sheet)
            new_sheet.title = template_sheet.title
            
            # 저장
            target_wb.save(target_file)
            target_wb.close()
            
            print(f"  ✓ 완료\n")
            success_count += 1
            
        except Exception as e:
            print(f"  ✗ 실패: {str(e)}\n")
            fail_count += 1
    
    # 결과 출력
    print("=" * 50)
    print(f"작업 완료!")
    print(f"성공: {success_count}개")
    print(f"실패: {fail_count}개")
    print("=" * 50)
    
    list_wb.close()
    template_wb.close()


# 사용 예시
if __name__ == "__main__":
    # 파일 경로 설정
    template_file = "템플릿.xlsx"  # 템플릿 시트가 있는 파일
    file_list_excel = "파일목록.xlsx"  # A열에 절대경로가 있는 파일
    
    # 실행
    copy_sheet_to_files(
        template_file=template_file,
        file_list_excel=file_list_excel,
        template_sheet_name=None  # None이면 첫 번째 시트
    )
